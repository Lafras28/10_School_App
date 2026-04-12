import csv
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

from attendance_register import get_daily_register, list_attendance_entries_for_range, update_attendance_entry
from incident_register import create_incident_record, list_incidents
from medicine_log import create_medicine_record, list_medicine_logs
from pdf_export import build_compliance_report_pdf

app = Flask(__name__)
CORS(app)

BASE_DIR = Path(__file__).resolve().parent.parent
STUDENTS_XLSX_FILE = BASE_DIR / 'Students' / 'students_template.xlsx'
STUDENTS_CSV_FILE = BASE_DIR / 'Students' / 'students_template.csv'
FALLBACK_XLSX_FILE = BASE_DIR / 'data' / 'students_template.xlsx'
FALLBACK_CSV_FILE = BASE_DIR / 'data' / 'students_template.csv'
STUDENT_COLUMNS = [
    'id',
    'firstName',
    'lastName',
    'emergencyContact1Name',
    'emergencyContact1Number',
    'emergencyContact2Name',
    'emergencyContact2Number',
    'emergencyContact3Name',
    'emergencyContact3Number',
    'allergies',
    'medicalAidName',
    'medicalAidNumber',
    'doctorContact',
    'medicalPin',
    'className',
]
DEFAULT_CLASSROOMS = ['Sunshine Bunnies', 'Rainbow Cubs', 'Little Explorers']


def infer_class_name(student_id):
    value = str(student_id or '').strip().lower()
    if value.startswith('st-') and value[3:].isdigit():
        return DEFAULT_CLASSROOMS[(int(value[3:]) - 1) % len(DEFAULT_CLASSROOMS)]

    return DEFAULT_CLASSROOMS[0]


def parse_parent_contacts(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return []

    # Support semicolon, comma, or slash-separated values from spreadsheets.
    separators = [';', ',', '/']
    parts = [value]
    for separator in separators:
        if separator in value:
            parts = [part.strip() for part in value.split(separator)]
            break

    return [part for part in parts if part]


def parse_emergency_contacts(row):
    contacts = []

    # Preferred format from API payload.
    payload_contacts = row.get('emergencyContacts')
    if isinstance(payload_contacts, list):
        for index, contact in enumerate(payload_contacts, start=1):
            if not isinstance(contact, dict):
                continue

            name = str(contact.get('name') or '').strip()
            number = str(contact.get('number') or '').strip()
            if not name and not number:
                continue

            contacts.append({
                'name': name or f'Emergency Contact {index}',
                'number': number,
            })

    # Spreadsheet columns format.
    if not contacts:
        for index in range(1, 4):
            name = str(row.get(f'emergencyContact{index}Name') or '').strip()
            number = str(row.get(f'emergencyContact{index}Number') or '').strip()
            if not name and not number:
                continue

            contacts.append({
                'name': name or f'Emergency Contact {index}',
                'number': number,
            })

    # Legacy fallback: parentContact string/list.
    if not contacts:
        parent_contact_raw = row.get('parentContact')
        if isinstance(parent_contact_raw, list):
            parsed_numbers = [str(number).strip() for number in parent_contact_raw if str(number).strip()]
        else:
            parsed_numbers = parse_parent_contacts(parent_contact_raw)

        for index, number in enumerate(parsed_numbers[:3], start=1):
            contacts.append({
                'name': f'Emergency Contact {index}',
                'number': number,
            })

    return contacts[:3]


def normalize_student_record(row):
    student_id = str(row.get('id') or '').strip()
    class_name = str(row.get('className') or row.get('classroom') or '').strip()

    return {
        'id': student_id,
        'firstName': str(row.get('firstName') or '').strip(),
        'lastName': str(row.get('lastName') or '').strip(),
        'emergencyContacts': parse_emergency_contacts(row),
        'allergies': str(row.get('allergies') or 'None').strip(),
        'medicalAidName': str(row.get('medicalAidName') or '').strip(),
        'medicalAidNumber': str(row.get('medicalAidNumber') or '').strip(),
        'doctorContact': str(row.get('doctorContact') or '').strip(),
        'medicalPin': str(row.get('medicalPin') or '').strip(),
        'className': class_name or infer_class_name(student_id),
    }


def students_to_sheet_row(student):
    contacts = student.get('emergencyContacts', [])

    def contact_name(position):
        return contacts[position - 1].get('name', '') if len(contacts) >= position else ''

    def contact_number(position):
        return contacts[position - 1].get('number', '') if len(contacts) >= position else ''

    return [
        student.get('id', ''),
        student.get('firstName', ''),
        student.get('lastName', ''),
        contact_name(1),
        contact_number(1),
        contact_name(2),
        contact_number(2),
        contact_name(3),
        contact_number(3),
        student.get('allergies', 'None'),
        student.get('medicalAidName', ''),
        student.get('medicalAidNumber', ''),
        student.get('doctorContact', ''),
        student.get('medicalPin', ''),
        student.get('className', infer_class_name(student.get('id', ''))),
    ]


def load_students_from_xlsx():
    target_file = STUDENTS_XLSX_FILE if STUDENTS_XLSX_FILE.exists() else FALLBACK_XLSX_FILE
    if not target_file.exists():
        return []

    workbook = load_workbook(target_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header).strip() if header is not None else '' for header in rows[0]]
    students = []

    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else '' for index in range(len(headers))}
        student = normalize_student_record(row)

        if student['id'] or student['firstName'] or student['lastName']:
            students.append(student)

    return students


def load_students_from_csv():
    target_file = STUDENTS_CSV_FILE if STUDENTS_CSV_FILE.exists() else FALLBACK_CSV_FILE
    if not target_file.exists():
        return []

    students = []
    with target_file.open(mode='r', encoding='utf-8', newline='') as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            students.append(normalize_student_record(row))

    return students


def load_students():
    students_from_xlsx = load_students_from_xlsx()
    if students_from_xlsx:
        return students_from_xlsx

    return load_students_from_csv()


def save_students_to_xlsx(students):
    if STUDENTS_XLSX_FILE.exists():
        workbook = load_workbook(STUDENTS_XLSX_FILE)
    elif FALLBACK_XLSX_FILE.exists():
        workbook = load_workbook(FALLBACK_XLSX_FILE)
    else:
        workbook = Workbook()

    sheet = workbook.active

    # Rewrite the worksheet from scratch to avoid stale rows after edits.
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(STUDENT_COLUMNS)

    for student in students:
        sheet.append(students_to_sheet_row(student))

    STUDENTS_XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(STUDENTS_XLSX_FILE)


def validate_student_payload(payload):
    missing = []

    for field in ['firstName', 'lastName']:
        if not str(payload.get(field, '')).strip():
            missing.append(field)

    contacts = payload.get('emergencyContacts', [])
    first_contact = contacts[0] if contacts else {}
    if not str(first_contact.get('name', '')).strip():
        missing.append('emergencyContact1Name')
    if not str(first_contact.get('number', '')).strip():
        missing.append('emergencyContact1Number')

    return missing


def next_student_id(students):
    highest = 0
    for student in students:
        value = str(student.get('id', '')).strip().lower()
        if value.startswith('st-'):
            suffix = value[3:]
            if suffix.isdigit():
                highest = max(highest, int(suffix))

    return f'st-{highest + 1:03d}'


def get_student_by_id(student_id):
    target_id = str(student_id or '').strip()
    if not target_id:
        return None

    return next((student for student in load_students() if student.get('id') == target_id), None)


def normalize_register_date(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return datetime.now().date().isoformat()

    try:
        return date.fromisoformat(value).isoformat()
    except ValueError as exc:
        raise ValueError('Date must use YYYY-MM-DD format.') from exc


def parse_created_at(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return None

    try:
        return datetime.fromisoformat(value.replace('Z', '+00:00'))
    except ValueError:
        return None


def filter_records_by_created_at(records, start_date, end_date, field_name='createdAt'):
    filtered = []
    for record in records:
        created_at = parse_created_at(record.get(field_name) or record.get('createdAt'))
        if created_at is None:
            continue

        created_date = created_at.date().isoformat()
        if start_date <= created_date <= end_date:
            filtered.append(record)

    return filtered


@app.get('/health')
def health_check():
    return jsonify({
        'status': 'ok',
        'service': 'school-safety-api',
        'timestamp': datetime.now(timezone.utc).isoformat(),
    }), 200


@app.post('/incident')
def create_incident():
    """
    Temporary endpoint that validates and accepts incident payloads.
    In production, this should write to Firestore via the Firebase Admin SDK.
    """
    payload = request.get_json(silent=True) or {}

    required_fields = ['schoolId', 'reportedBy', 'incidentType', 'description']
    missing = [field for field in required_fields if not payload.get(field)]

    if missing:
        return jsonify({
            'error': 'Missing required fields.',
            'missingFields': missing,
        }), 400

    response = {
        'message': 'Incident received.',
        'incident': {
            'schoolId': payload['schoolId'],
            'reportedBy': payload['reportedBy'],
            'incidentType': payload['incidentType'],
            'description': payload['description'],
            'childId': payload.get('childId'),
            'severity': payload.get('severity', 'unspecified'),
            'createdAt': datetime.now(timezone.utc).isoformat(),
        },
    }

    return jsonify(response), 201


@app.get('/attendance')
def list_attendance():
    try:
        register_date = normalize_register_date(request.args.get('date'))
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    students = load_students()
    entries = get_daily_register(register_date, students)
    return jsonify({'date': register_date, 'entries': entries}), 200


@app.put('/attendance/<register_date>/<student_id>')
def update_attendance(register_date, student_id):
    student = get_student_by_id(student_id)
    if student is None:
        return jsonify({'error': 'Student not found.'}), 404

    payload = request.get_json(silent=True) or {}
    try:
        normalized_date = normalize_register_date(register_date)
        entry = update_attendance_entry(
            normalized_date,
            student,
            payload.get('status', 'Present'),
            payload.get('reason', ''),
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    return jsonify({'message': 'Attendance updated.', 'entry': entry}), 200


@app.get('/incidents')
def list_incident_records():
    return jsonify(list_incidents()), 200


@app.post('/incidents')
def create_incident_entry():
    payload = request.get_json(silent=True) or {}
    student_id = str(payload.get('studentId') or '').strip()
    student = get_student_by_id(student_id) if student_id else None

    if student_id and student is None:
        return jsonify({'error': 'Student not found.'}), 404

    try:
        incident = create_incident_record(payload, student)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    return jsonify({'message': 'Incident recorded and locked.', 'incident': incident}), 201


@app.route('/incidents/<incident_id>', methods=['PUT', 'PATCH', 'DELETE'])
def incident_record_read_only(incident_id):
    return jsonify({
        'error': 'Incident records are read-only after saving for legal compliance.',
        'incidentId': incident_id,
    }), 403


@app.get('/medicine')
def list_medicine_entries():
    return jsonify(list_medicine_logs()), 200


@app.post('/medicine')
def create_medicine_entry():
    payload = request.get_json(silent=True) or {}
    student_id = str(payload.get('studentId') or '').strip()
    student = get_student_by_id(student_id)
    if student is None:
        return jsonify({'error': 'Student not found.'}), 404

    try:
        entry = create_medicine_record(payload, student)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    return jsonify({
        'message': 'Medicine administration logged.',
        'entry': entry,
        'warning': 'WARNING: Medication matches a recorded allergy.' if entry.get('allergyWarning') else '',
    }), 201


@app.get('/exports/compliance-report')
def export_compliance_report():
    school_name = str(request.args.get('schoolName') or 'School Safety & Compliance Centre').strip()
    try:
        start_date = normalize_register_date(request.args.get('startDate'))
        end_date = normalize_register_date(request.args.get('endDate') or start_date)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    if start_date > end_date:
        return jsonify({'error': 'startDate must be before or equal to endDate.'}), 400

    students = load_students()
    attendance_entries = list_attendance_entries_for_range(start_date, end_date, students)
    # Filter to only include Late and Absent entries, exclude Present
    attendance_entries = [entry for entry in attendance_entries if entry.get('status') in ['Late', 'Absent']]
    incidents = filter_records_by_created_at(list_incidents(), start_date, end_date, field_name='timestamp')
    medicine_entries = filter_records_by_created_at(list_medicine_logs(), start_date, end_date, field_name='timeAdministered')

    pdf_bytes = build_compliance_report_pdf(
        school_name=school_name,
        start_date=start_date,
        end_date=end_date,
        attendance_entries=attendance_entries,
        incidents=incidents,
        medicine_logs=medicine_entries,
    )

    filename = f'compliance-report-{start_date}-to-{end_date}.pdf'
    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename,
    )


@app.get('/students')
def list_students():
    return jsonify(load_students()), 200


@app.post('/students')
def create_student():
    payload = request.get_json(silent=True) or {}
    missing = validate_student_payload(payload)
    if missing:
        return jsonify({'error': 'Missing required fields.', 'missingFields': missing}), 400

    students = load_students()
    new_student = normalize_student_record(payload)
    if not new_student['id']:
        new_student['id'] = next_student_id(students)

    existing = next((student for student in students if student['id'] == new_student['id']), None)
    if existing:
        return jsonify({'error': 'Student id already exists.'}), 409

    students.append(new_student)
    save_students_to_xlsx(students)
    return jsonify({'message': 'Student added.', 'student': new_student}), 201


@app.put('/students/<student_id>')
def update_student(student_id):
    payload = request.get_json(silent=True) or {}
    students = load_students()

    index = next((idx for idx, student in enumerate(students) if student['id'] == student_id), None)
    if index is None:
        return jsonify({'error': 'Student not found.'}), 404

    merged_student = {
        **students[index],
        **normalize_student_record({**students[index], **payload, 'id': student_id}),
    }

    missing = validate_student_payload(merged_student)
    if missing:
        return jsonify({'error': 'Missing required fields.', 'missingFields': missing}), 400

    students[index] = merged_student
    save_students_to_xlsx(students)
    return jsonify({'message': 'Student updated.', 'student': merged_student}), 200


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
